from jira import JIRA
from openpyxl import Workbook, load_workbook

username = ''
password = ''

# dictionary of custom project fields
project_dict = {
    'PROJECT': 'customfield_1',
}
issutype_dict ={
    'PROJECT': 'Task',
}

def create_comment_with_mention(issue_key, mentioned_user):
    issue = jira.issue(issue_key)
    comment = issue.fields.comment
    comment_text = ', Please look at the issue'
    comment_text_with_mention =f'{mentioned_user} {comment_text}'
    jira.add_comment(issue, body=comment_text_with_mention)

jira = JIRA(server='', basic_auth=(username, password))

excel_book = load_workbook('jira.xlsx')
sheets = excel_book.sheetnames
ws = excel_book[sheets[0]]

end_of_file = ws.max_row
end_of_file += 1

# send to create jira task
j = 2
while j < end_of_file:
    cell = 'a' + str(j)
    project_name = ws[cell].value

    cell5 = 'f' + str(j)
    root_issue_key = ws[cell5].value
    root_issue = jira.issue(root_issue_key)

    cell7 = 'i' + str(j)
    duedate = ws[cell7].value
    if duedate != None:
        duedate = str(duedate)
        duedate = duedate[:10]

    cell8 = 'h' + str(j)
    group = ws[cell8].value

    cell9 = 'j' + str(j)
    mentioned_user = ws[cell9].value


    fields = {
        'project': str(project_name),
        'summary': str(root_issue.fields.summary),
        'description': str(root_issue.fields.description),
        'priority': {'id': str(root_issue.fields.priority.id)}
    }

    if project_name in issutype_dict:
        custom_issuetype_name = project_dict[custom_issuetype_name]
        fields['issuetype'] = {'name': custom_issuetype_name}
    else:
        fields['issuetype'] = {'name': 'Story'}

    if group != None:
        if project_name in project_dict:
            custom_field_name = project_dict[project_name]
            fields[custom_field_name] = {'value': group}

    if duedate != None:
        fields['duedate'] = duedate

    try:
        clone_issue = jira.create_issue(fields=fields)

        link = jira.create_issue_link(
            type='issue for fix',
            inwardIssue=str(clone_issue),
            outwardIssue=str(root_issue))

        cell6 = 'g' + str(j)
        clone_issue = str(clone_issue)
        ws[cell6].value = clone_issue

        if mentioned_user != None:
            mentioned_user = '[~' + mentioned_user + ']'
            create_comment_with_mention(clone_issue, mentioned_user)

    except Exception as e:
        cell6 = 'g' + str(j)
        clone_issue = 'Task not created'
        ws[cell6].value = clone_issue
        print(e)
    j += 1

excel_book.save('jira.xlsx')
excel_book.close()
