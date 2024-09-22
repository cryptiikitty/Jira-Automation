from __future__ import annotations
import time
from jira import JIRA
from openpyxl import Workbook, load_workbook


username = ''
password = ''


try:
    jira = JIRA(server='', basic_auth=(username, password))

    excel_book = load_workbook('jira.xlsx')
    sheets = excel_book.sheetnames
    ws = excel_book[sheets[0]]

    end_of_file = ws.max_row
    end_of_file += 1
    j = 2

    # send to create jira task
    j = 2
    while j < end_of_file:
        time.sleep(5)
        cell = 'a' + str(j)
        project_name = ws[cell].value

        cell2 = 'b' + str(j)
        summ = ws[cell2].value


        cell3 = 'c' + str(j)
        prior = ws[cell3].value

        cell4 = 'e' + str(j)
        desc = ws[cell4].value

        cell6 = 'd' + str(j)
        category = ws[cell6].value
        # Specify the project in which tasks will be created
        new_issue = jira.create_issue(project='PROJECT', summary= str(summ),
                                      description=str(desc), issuetype={'name': 'Story'},
                                      priority={'id': str(prior)}, customfield_11243={'value': str(category)})

        cell5 = 'f' + str(j)
        new_issue = str(new_issue)
        ws[cell5].value = new_issue

        j += 1
    excel_book.save('jira.xlsx')
    excel_book.close()


except:
    print('Error')

